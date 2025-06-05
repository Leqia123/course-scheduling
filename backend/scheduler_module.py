# scheduler_module.py
# -*- coding: utf-8 -*-
import psycopg2
import psycopg2.extras
import datetime
import math
import random
import copy
import pandas as pd
from collections import defaultdict, namedtuple
import re
import io

# --- 检查 openpyxl 库 ---
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # print("警告：未找到 'openpyxl' 库，将无法导出 Excel 课表。请运行 'pip install openpyxl' 安装。") # 在app.py中处理

# ==================================
# 2. 数据模型 (保持不变)
# ==================================
# ... (你的数据模型定义保持不变) ...
Semester = namedtuple('Semester', ['id', 'name', 'start_date', 'end_date', 'total_weeks'])
Major = namedtuple('Major', ['id', 'name'])
Teacher = namedtuple('Teacher', ['id', 'user_id', 'name'])
Classroom = namedtuple('Classroom', ['id', 'name', 'capacity', 'type'])
Course = namedtuple('Course', ['id', 'name', 'total_sessions', 'course_type'])
TimeSlot = namedtuple('TimeSlot', ['id', 'day_of_week', 'period', 'start_time', 'end_time'])
CourseAssignment = namedtuple('CourseAssignment',
                              ['id', 'major_id', 'course_id', 'teacher_id', 'semester_id', 'is_core_course',
                               'expected_students'])
TimetableEntry = namedtuple('TimetableEntry',
                            ['id', 'semester_id', 'major_id', 'course_id', 'teacher_id', 'classroom_id', 'timeslot_id',
                             'week_number', 'assignment_id'])
# --- 新增：教师偏好数据结构 ---
TeacherPreference = namedtuple('TeacherPreference', ['id', 'teacher_id', 'semester_id', 'timeslot_id', 'preference_type', 'status', 'reason'])
# --- 新增结束 ---


# ==================================
# 3. 数据加载函数 (修改: 加载教师偏好)
# ==================================
def load_data_from_db(get_connection_func):
    """从数据库加载所有基础数据，包括教师偏好"""
    print("SCHEDULER: 开始从数据库加载数据...")
    all_data = {}
    conn = None
    cur = None # Define cur outside try
    try:
        conn = get_connection_func()  # 使用传入的函数获取连接
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        # ... (其余加载逻辑与您脚本中的 load_data_from_db 基本相同) ...
        # 加载学期 (计算 total_weeks)
        cur.execute("SELECT id, name, start_date, end_date FROM semesters")
        raw_semesters = cur.fetchall()
        all_data['semesters'] = {}
        # print("  - 正在加载学期信息并计算总周数...")
        for row in raw_semesters:
            start_date = row['start_date']
            end_date = row['end_date']
            calculated_weeks = 0
            if isinstance(start_date, datetime.date) and isinstance(end_date, datetime.date):
                if end_date >= start_date:
                    delta_days = (end_date - start_date).days + 1
                    calculated_weeks = math.ceil(delta_days / 7)
            all_data['semesters'][row['id']] = Semester(
                id=row['id'], name=row['name'], start_date=start_date,
                end_date=end_date, total_weeks=calculated_weeks
            )
        # print(f"  - 加载并处理了 {len(all_data['semesters'])} 个学期信息")

        # 加载专业
        cur.execute("SELECT id, name FROM majors")
        all_data['majors'] = {row['id']: Major(**row) for row in cur.fetchall()}
        # print(f"  - 加载了 {len(all_data['majors'])} 个专业信息")

        # 加载教师
        cur.execute("SELECT id, username FROM users")
        user_id_to_name = {row['id']: row['username'] for row in cur.fetchall()}
        cur.execute("SELECT id, user_id FROM teachers")
        raw_teachers = cur.fetchall()
        all_data['teachers'] = {}
        for row in raw_teachers:
            teacher_id = row['id']
            user_id = row['user_id']
            teacher_name = user_id_to_name.get(user_id, f"未知用户(ID:{user_id})")
            all_data['teachers'][teacher_id] = Teacher(id=teacher_id, user_id=user_id, name=teacher_name)
        # print(f"  - 加载并处理了 {len(all_data['teachers'])} 个教师信息")

        # 加载教室
        cur.execute("SELECT id, building, room_number, capacity, room_type FROM classrooms")
        all_data['classrooms'] = {}
        for row in cur.fetchall():
            classroom_id = row['id']
            building_name = row['building'] if row['building'] else '未知楼'
            room_num = row['room_number'] if row['room_number'] else '未知号'
            classroom_name = f"{building_name}-{room_num}"
            all_data['classrooms'][classroom_id] = Classroom(
                id=classroom_id, name=classroom_name, capacity=row['capacity'], type=row['room_type']
            )
        # print(f"  - 加载了 {len(all_data['classrooms'])} 个教室信息")

        # 加载课程
        cur.execute("SELECT id, name, total_sessions, course_type FROM courses")
        all_data['courses'] = {row['id']: Course(**row) for row in cur.fetchall()}
        # print(f"  - 加载了 {len(all_data['courses'])} 个课程信息")

        # 加载时间段
        cur.execute("SELECT id, day_of_week, period, start_time, end_time FROM time_slots ORDER BY day_of_week, period")
        all_data['timeslots'] = {row['id']: TimeSlot(**row) for row in cur.fetchall()}
        all_data['timeslot_lookup'] = {(ts.day_of_week, ts.period): ts.id for ts in all_data['timeslots'].values()}
        # print(f"  - 加载了 {len(all_data['timeslots'])} 个时间段信息")

        # 加载教学任务
        cur.execute("""
            SELECT id, major_id, course_id, teacher_id, semester_id, is_core_course, expected_students
            FROM course_assignments
        """)
        all_data['course_assignments'] = {row['id']: CourseAssignment(**row) for row in cur.fetchall()}
        # print(f"  - 加载了 {len(all_data['course_assignments'])} 个教学任务")

        # --- 新增：加载教师排课偏好 ---
        cur.execute("""
            SELECT id, teacher_id, semester_id, timeslot_id, preference_type, status, reason
            FROM teacher_scheduling_preferences
            WHERE  preference_type = 'avoid'
        """)
        # Store approved 'avoid' preferences in a set for quick lookup: (teacher_id, timeslot_id, semester_id)
        all_data['approved_avoid_preferences'] = set()
        raw_preferences = cur.fetchall()
        for row in raw_preferences:
             all_data['approved_avoid_preferences'].add((row['teacher_id'], row['timeslot_id'], row['semester_id']))
        # print(f"  - 加载并过滤了 {len(raw_preferences)} 条教师偏好记录 (其中 {len(all_data['approved_avoid_preferences'])} 条已批准的'避免'偏好将作为约束)")
        # --- 新增结束 ---

        cur.close()
        print("SCHEDULER: 数据加载成功!")
        return all_data

    except psycopg2.Error as e:
        print(f"SCHEDULER: 数据库连接或查询错误: {e}")
        # import traceback; traceback.print_exc() # For more detailed server-side logs if needed
        raise  # Re-raise the exception to be caught by the Flask route
    except Exception as e:
        print(f"SCHEDULER: 数据加载过程中发生未知错误: {e}")
        # import traceback; traceback.print_exc()
        raise # Re-raise
    finally:
        if cur: cur.close()
        if conn: conn.close()


# ==================================
# 4. 辅助函数 (修改: check_constraints 增加偏好检查)
# ==================================
# ... (你的 find_timeslot_id 保持不变) ...
def find_timeslot_id(day_str, period_num, all_data):
    return all_data['timeslot_lookup'].get((day_str, period_num))


# 注意：check_constraints 需要使用传入的全局状态和 all_data
# 修改：返回一个表示是否成功的布尔值和一个表示冲突原因的字符串（如果失败）
def check_constraints(timetable_state, assignment, week, timeslot_id, classroom_id, all_data):
    teacher_id = assignment.teacher_id
    major_id = assignment.major_id
    semester_id = assignment.semester_id # 获取学期 ID

    # --- 新增：检查教师的“避免安排”偏好 ---
    if (teacher_id, timeslot_id, semester_id) in all_data.get('approved_avoid_preferences', set()):
        # print(f"[CONFLICT] Teacher {teacher_id} has 'avoid' preference for timeslot {timeslot_id} in semester {semester_id}")
        return False, "教师偏好 (避免安排)"
    # --- 新增结束 ---

    # 检查全局状态中教师、教室、专业是否已被占用
    if (teacher_id, week, timeslot_id) in timetable_state['teacher_schedule']:
        # print(f"[CONFLICT] Teacher {teacher_id} busy week {week} slot {timeslot_id}")
        return False, "教师冲突 (已安排其它课程)"
    if (classroom_id, week, timeslot_id) in timetable_state['classroom_schedule']:
        # print(f"[CONFLICT] Classroom {classroom_id} busy week {week} slot {timeslot_id}")
        return False, "教室冲突 (已被占用)"
    if (major_id, week, timeslot_id) in timetable_state['major_schedule']:
        # print(f"[CONFLICT] Major {major_id} busy week {week} slot {timeslot_id}")
        return False, "专业冲突 (已安排其它课程)"

    return True, None # 没有冲突

# 注意：find_available_classroom 需要使用传入的全局状态 (保持不变)
# ... (你的 find_available_classroom 函数保持不变) ...
def find_available_classroom(timetable_state, assignment, week, timeslot_id, all_data):
    required_capacity = assignment.expected_students
    course = all_data['courses'].get(assignment.course_id)
    is_lab_course = course and course.course_type == '实验课'
    preferred_type_available, other_type_available = [], []

    # 确定在当前周次和时间段已经被占用的教室集合，检查全局状态
    busy_classrooms = {cid for (cid, w, tid) in timetable_state['classroom_schedule'] if
                       w == week and tid == timeslot_id}

    for classroom_id, classroom in all_data['classrooms'].items():
        # 检查教室是否在全局状态中已被占用
        if classroom_id in busy_classrooms: continue
        if classroom.capacity < required_capacity: continue
        classroom_type_match = (is_lab_course and classroom.type == '实验室') or \
                               (not is_lab_course and classroom.type == '普通教室')
        if classroom_type_match:
            preferred_type_available.append(classroom_id)
        else:
            other_type_available.append(classroom_id)

    if preferred_type_available: return random.choice(preferred_type_available)
    if other_type_available: return random.choice(other_type_available)
    return None


# ==================================
# 5. 自动生成初始模板函数 (保持不变)
# ==================================
# ... (你的 generate_initial_template 函数保持不变) ...
def generate_initial_template(assignments_dict, all_data):
    # print("SCHEDULER:   正在根据可用任务自动生成初始周模板...")
    day_order = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    template_structure = [
        ('周一', 1, '理论课', 'MonWed1'), ('周一', 2, '理论课', 'MonWed2'), ('周一', 3, None, None),
        ('周一', 4, None, None),
        ('周二', 1, '理论课', 'TueThu1'), ('周二', 2, '理论课', 'TueThu2'), ('周二', 3, None, None),
        ('周二', 4, None, None),
        ('周三', 1, '理论课', 'MonWed1'), ('周三', 2, '理论课', 'MonWed2'), ('周三', 3, None, None),
        ('周三', 4, None, None),
        ('周四', 1, '理论课', 'TueThu1'), ('周四', 2, '理论课', 'TueThu2'), ('周四', 3, None, None),
        ('周四', 4, None, None),
        ('周五', 1, '实验课', 'FriLab1'), ('周五', 2, '实验课', 'FriLab2'), ('周五', 3, None, None),
        ('周五', 4, None, None),
        ('周六', 1, None, None), ('周六', 2, None, None), ('周六', 3, None, None), ('周六', 4, None, None),
        ('周日', 1, None, None), ('周日', 2, None, None), ('周日', 3, None, None), ('周日', 4, None, None),
    ] # 包含周六周日，即使默认不排课，但时间段可能存在

    repetition_map = defaultdict(list)
    slot_type_map = {}
    initial_template = {}
    all_slots_in_template = set()

    # 构建基于数据库 timeslots 的模板结构，而不是硬编码
    db_timeslots_raw = sorted(all_data['timeslots'].values(), key=lambda t: (t.day_of_week, t.period))
    db_template_structure = [(ts.day_of_week, ts.period, None, None) for ts in db_timeslots_raw] # Default type None, no groups
     # You could potentially enhance this to read 'preferred' types or groups from configuration or DB
     # For now, let's stick closer to the original but use DB timeslots
    # Let's map time slot IDs to (day, period) for easier lookup against the template structure
    id_to_day_period = {ts.id: (ts.day_of_week, ts.period) for ts in all_data['timeslots'].values()}
    # Rebuild template structure using actual timeslot IDs
    initial_template_by_id = {ts_id: None for ts_id in all_data['timeslots'].keys()}
    # We still need the (day, period) for sorting and potentially basic grouping/type matching
    sorted_timeslot_ids = sorted(all_data['timeslots'].keys(),
                                 key=lambda ts_id: (day_order.index(all_data['timeslots'][ts_id].day_of_week) if all_data['timeslots'][ts_id].day_of_week in day_order else 99,
                                                    all_data['timeslots'][ts_id].period))


    # Original grouping logic (can be adapted or replaced)
    # For simplicity, let's revert to the original template structure logic for generating the *initial* template fill,
    # but ensure we use timeslot IDs correctly later. The template is more about *which* assignments *could* go where
    # in a recurring pattern, before weekly constraints are applied.
    # The *actual* scheduling uses the timeslot_id derived from (day_str, period_num).

    template_structure_simplified = [
        ('周一', 1), ('周一', 2), ('周一', 3), ('周一', 4),
        ('周二', 1), ('周二', 2), ('周二', 3), ('周二', 4),
        ('周三', 1), ('周三', 2), ('周三', 3), ('周三', 4),
        ('周四', 1), ('周四', 2), ('周四', 3), ('周四', 4),
        ('周五', 1), ('周五', 2), ('周五', 3), ('周五', 4),
        ('周六', 1), ('周六', 2), ('周六', 3), ('周六', 4),
        ('周日', 1), ('周日', 2), ('周日', 3), ('周日', 4),
    ]
    # Filter template structure to only include time slots that actually exist in the DB
    template_slots_from_db = [
        (day, period) for day, period in template_structure_simplified
        if (day, period) in all_data['timeslot_lookup']
    ]
    day_order = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    sorted_template_slots_dp = sorted(
        template_slots_from_db,
        key=lambda x: (day_order.index(x[0]) if x[0] in day_order else 99, x[1])
    )

    initial_template_fill = {} # Maps (day, period) to assignment_id
    all_template_slots_dp = set(sorted_template_slots_dp)

    # Let's use a simpler approach for initial fill: just assign tasks greedily to available slots
    # prioritizing core courses / more sessions.
    # This template fill doesn't handle repetitions like Mon/Wed grouping, simplifying the approach.
    # The weekly scheduling loop is where the main logic applies constraints.

    theory_assignments, lab_assignments, other_assignments = [], [], []
    if not assignments_dict: return {}, [] # Return empty template and pool

    def get_priority(assign_id, assign):
        course = all_data['courses'].get(assign.course_id)
        # Priority: is_core_course (True > False), total_sessions (more > less), then random
        return (assign.is_core_course, -(course.total_sessions if course else 0), random.random())

    # Sort all assignments by priority
    all_assignments_sorted = sorted(
         ((get_priority(assign_id, assign), assign_id) for assign_id, assign in assignments_dict.items()),
         reverse=True
    )
    assignment_pool_ids = [assign_id for _, assign_id in all_assignments_sorted]

    # Attempt to fill the initial template slots
    used_assignment_ids = set()
    template_slot_index = 0

    while assignment_pool_ids and template_slot_index < len(sorted_template_slots_dp):
        assign_id_to_fill = assignment_pool_ids.pop(0) # Take the highest priority assignment
        if assign_id_to_fill in used_assignment_ids: continue # Should not happen with pop(0) from pool
        assign = assignments_dict.get(assign_id_to_fill)
        if not assign: continue
        course = all_data['courses'].get(assign.course_id)
        if not course or course.total_sessions <= 0: continue

        current_slot_dp = sorted_template_slots_dp[template_slot_index]

        # Check if this assignment is suitable for this slot type (basic check)
        # This part could be expanded based on course types vs preferred time slot types
        # For now, any course can go into any slot in the *initial template* for simplicity.
        # The real constraint checks happen weekly.

        initial_template_fill[current_slot_dp] = assign_id_to_fill
        used_assignment_ids.add(assign_id_to_fill)
        template_slot_index += 1 # Move to the next template slot

    # The remaining assignments in the pool are those that didn't fit into the initial template slots
    # because the slots ran out or assignments were invalid.
    # However, assignments *in* the template might still have sessions remaining after filling.
    # A better unscheduled pool is simply all assignments that started with sessions > 0
    # The weekly loop will track remaining sessions.
    # So, the unscheduled pool is effectively all assignments initially needing scheduling.
    # Let's adjust the return value to just be the initial template fill map.
    # The 'unscheduled pool' concept is better handled by tracking sessions remaining *during* the weekly loop.

    # The original code's `unscheduled_pool_ids` was the list of assignments *not* put into the initial template.
    # Let's keep that concept for now, as the original logic used it for replacement.
    # Remaining assignments after initial template fill:
    unscheduled_pool_ids = [
        assign_id for assign_id in assignments_dict
        if assign_id not in used_assignment_ids
        and all_data['courses'].get(assignments_dict[assign_id].course_id, Course(None, None, 0, None)).total_sessions > 0
    ]
    # print(f"SCHEDULER:   自动生成模板完成。选入 {len(used_assignment_ids)} 个任务到模板。剩余 {len(unscheduled_pool_ids)} 个任务进入替换池。")

    # Return template based on (day, period) tuples
    return initial_template_fill, unscheduled_pool_ids


# ==================================
# 6. 基于模板的排课执行函数 (修改: 接收并更新全局状态，处理冲突原因)
# ==================================
def schedule_with_generated_template(assignments_for_major, current_semester, current_major, all_data, initial_template_dp, # initial_template is now (day, period) map
                                     unscheduled_pool_ids, global_timetable_state):
    # print(f"\nSCHEDULER: ===== 开始为专业 '{current_major.name}' 基于模板排课 (学期: {current_semester.name}, {current_semester.total_weeks} 周) =====")
    total_weeks = current_semester.total_weeks
    if not total_weeks or total_weeks <= 0:
        unscheduled_details_on_error = []
        for assign_id, assign in assignments_for_major.items():
            course = all_data['courses'].get(assign.course_id)
            teacher = all_data['teachers'].get(assign.teacher_id)
            unscheduled_details_on_error.append({
                'assignment_id': assign_id,
                'course_name': course.name if course else '未知课程',
                'teacher_name': teacher.name if teacher else '未知教师',
                'remaining_sessions': course.total_sessions if course else 0
            })
        return {'schedule': [], 'unscheduled_details': unscheduled_details_on_error, 'conflicts': []}

    assignment_sessions_remaining = {}
    for assign_id, assign in assignments_for_major.items():
        course = all_data['courses'].get(assign.course_id)
        assignment_sessions_remaining[assign_id] = course.total_sessions if course else 0

    unscheduled_pool = list(unscheduled_pool_ids)
    # current_template_schedule = initial_template_dp.copy() # This template is just an initial suggestion,
    # we should check assignment sessions every week

    final_schedule = []
    conflicts_log = []

    day_order = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    # Get actual timeslot IDs sorted by day and period
    sorted_timeslot_ids_by_dp = sorted(
        all_data['timeslots'].keys(),
        key=lambda ts_id: (day_order.index(all_data['timeslots'][ts_id].day_of_week) if all_data['timeslots'][ts_id].day_of_week in day_order else 99,
                           all_data['timeslots'][ts_id].period)
    )

    # We will iterate through weeks and then sorted time slots (by ID, which corresponds to day/period)
    # For each slot, we check the *initial template* to see which assignment *was suggested* for this (day, period)
    # Then check if that assignment still needs sessions. If so, try to schedule it.
    # If not, try to find a replacement from the unscheduled pool.

    # Pre-map timeslot_id to (day, period) for easier template lookup
    timeslot_id_to_dp = {ts.id: (ts.day_of_week, ts.period) for ts in all_data['timeslots'].values()}


    # Build a dynamic pool of assignments that still need scheduling, ordered by initial priority
    # Start with all assignments needing sessions
    dynamic_unscheduled_assignments = sorted(
        [assign_id for assign_id, rem_sessions in assignment_sessions_remaining.items() if rem_sessions > 0],
        key=lambda assign_id: (assignments_for_major.get(assign_id) and all_data['courses'].get(assignments_for_major[assign_id].course_id) and assignments_for_major[assign_id].is_core_course,
                               -(assignments_for_major.get(assign_id) and all_data['courses'].get(assignments_for_major[assign_id].course_id).total_sessions or 0),
                               random.random()),
        reverse=True
    )
    # Add assignments from the original unscheduled pool (those not in the initial template)
    # Ensure no duplicates and they still need sessions
    for assign_id in unscheduled_pool:
        if assign_id not in dynamic_unscheduled_assignments and assignment_sessions_remaining.get(assign_id, 0) > 0:
             # Maintain some order if possible, or shuffle slightly
             dynamic_unscheduled_assignments.append(assign_id)
    random.shuffle(dynamic_unscheduled_assignments) # Shuffle the dynamic pool to avoid infinite loops on conflicts


    for week in range(1, total_weeks + 1):
        # In each week, we can iterate through the available time slots
        # For each slot, determine which assignment to try scheduling
        # Option 1: Strictly follow the template if the task isn't finished and fits
        # Option 2: For slots where the template task is finished or failed, pick from the dynamic pool

        assignments_tried_this_week = set() # To avoid trying the same assignment multiple times in the same week if it fails

        for timeslot_id in sorted_timeslot_ids_by_dp:
            day_str, period_num = timeslot_id_to_dp.get(timeslot_id, (None, None))
            if day_str is None: continue # Should not happen if timeslot_id is valid

            # Determine which assignment to attempt for this slot/week
            assignment_to_attempt_id = None

            # Check the initial template suggestion for this slot (day, period)
            suggested_assign_id = initial_template_dp.get((day_str, period_num))

            if suggested_assign_id is not None and assignment_sessions_remaining.get(suggested_assign_id, 0) > 0:
                 # If the suggested assignment still needs sessions, try it first
                 assignment_to_attempt_id = suggested_assign_id
                 # print(f"  Week {week}, {day_str}-{period_num}: Trying template suggested task {suggested_assign_id}")
            else:
                 # If template task is finished or invalid, try from the dynamic unscheduled pool
                 # Find the next assignment from the pool that still needs sessions
                 found_in_pool = False
                 for assign_id_from_pool in list(dynamic_unscheduled_assignments): # Iterate over a copy
                      if assignment_sessions_remaining.get(assign_id_from_pool, 0) > 0:
                           # Check if trying this assignment for this specific slot (timeslot_id)
                           # would immediately violate the teacher's 'avoid' preference
                           # Do a quick check here BEFORE calling check_constraints to potentially skip early
                           assign_from_pool = assignments_for_major.get(assign_id_from_pool)
                           if assign_from_pool and (assign_from_pool.teacher_id, timeslot_id, assign_from_pool.semester_id) in all_data.get('approved_avoid_preferences', set()):
                                # This assignment cannot be scheduled in this slot due to teacher preference
                                # print(f"  Week {week}, {day_str}-{period_num}: Skipping pool task {assign_id_from_pool} due to teacher preference conflict.")
                                continue # Try next assignment in the pool

                           # If it passed the preference check (or preference check is not the first step in check_constraints),
                           # select this one. Remove it from the pool for this attempt.
                           assignment_to_attempt_id = assign_id_from_pool
                           dynamic_unscheduled_assignments.remove(assign_id_from_pool) # Remove from dynamic pool
                           found_in_pool = True
                           # print(f"  Week {week}, {day_str}-{period_num}: Trying pool task {assign_id_from_pool}")
                           break # Found an assignment to attempt from the pool

                 if not found_in_pool:
                     # No suitable assignment found in the pool that still needs sessions and passes quick checks
                     # print(f"  Week {week}, {day_str}-{period_num}: No task needing sessions or suitable found from pool.")
                     continue # Move to the next timeslot


            # Now try to schedule `assignment_to_attempt_id` in this `(week, timeslot_id)`
            if assignment_to_attempt_id is not None and assignment_sessions_remaining.get(assignment_to_attempt_id, 0) > 0:
                 assignment = assignments_for_major.get(assignment_to_attempt_id)
                 if not assignment: continue # Should not happen

                 # Find a classroom
                 suitable_classroom_id = find_available_classroom(global_timetable_state, assignment, week, timeslot_id, all_data)

                 if suitable_classroom_id:
                     # Check all constraints (including the new preference check inside)
                     is_possible, conflict_reason = check_constraints(global_timetable_state, assignment, week, timeslot_id, suitable_classroom_id, all_data)

                     if is_possible:
                         # Success! Create timetable entry
                         entry = TimetableEntry(None, current_semester.id, assignment.major_id, assignment.course_id,
                                                assignment.teacher_id, suitable_classroom_id, timeslot_id, week,
                                                assignment_to_attempt_id)
                         final_schedule.append(entry)

                         # Update global state
                         global_timetable_state['teacher_schedule'].add((assignment.teacher_id, week, timeslot_id))
                         global_timetable_state['classroom_schedule'].add((suitable_classroom_id, week, timeslot_id))
                         global_timetable_state['major_schedule'].add((assignment.major_id, week, timeslot_id))

                         # Decrement remaining sessions
                         assignment_sessions_remaining[assignment_to_attempt_id] -= 1
                         # print(f"  SUCCESS: Week {week}, {day_str}-{period_num} assigned {assignment_to_attempt_id}. Remaining: {assignment_sessions_remaining[assignment_to_attempt_id]}")

                         # If task is finished, remove it from the dynamic pool
                         if assignment_sessions_remaining[assignment_to_attempt_id] == 0 and assignment_to_attempt_id in dynamic_unscheduled_assignments:
                              try:
                                  dynamic_unscheduled_assignments.remove(assignment_to_attempt_id)
                                  # print(f"  TASK COMPLETED: {assignment_to_attempt_id} removed from dynamic pool.")
                              except ValueError:
                                  pass # Already removed or not in pool


                     else:
                         # Constraint conflict, log it
                         conflicts_log.append(
                             {'major_id': assignment.major_id, 'week': week, 'day': day_str, 'period': period_num,
                              'assignment_id': assignment_to_attempt_id, 'reason': conflict_reason})
                         # print(f"  CONFLICT: Week {week}, {day_str}-{period_num} attempt {assignment_to_attempt_id} failed: {conflict_reason}")

                         # If the attempted assignment came from the dynamic pool and failed, put it back
                         # so it can be attempted in a later slot/week.
                         # IMPORTANT: If it was the suggested template task, it *doesn't* go back to the *dynamic* pool
                         # unless it was also in the initial unscheduled pool.
                         # Let's simplify: if a task from the *dynamic_unscheduled_assignments* list fails, put it back.
                         # If it was from the *initial template fill* and NOT in the original unscheduled_pool,
                         # it was a 'template-only' suggestion. If it failed the *first time* it was tried,
                         # maybe it should go into the dynamic pool?
                         # Let's stick to the rule: only put back if it came from `dynamic_unscheduled_assignments`.
                         # The initial template fill is just a *first attempt* order. If it fails,
                         # it effectively becomes an unscheduled task.
                         # So, if the task came from `dynamic_unscheduled_assignments`, put it back.
                         # If it came from `initial_template_dp` and was *not* in the original `unscheduled_pool_ids`,
                         # consider adding it to the `dynamic_unscheduled_assignments` pool.
                         # This logic is getting complex. A simpler approach is to say: Any task needing sessions
                         # is potentially schedulable. If it fails *this* slot/week, it remains needing sessions
                         # and will be picked from the dynamic pool for a *later* slot/week if possible.
                         # So, if attempt failed, do *not* decrement sessions, and the task remains in the pool (implicitly)
                         # or needs to be added back if it was taken from pool.

                         # Let's refine: If `assignment_to_attempt_id` was taken from `dynamic_unscheduled_assignments`
                         # for this attempt, and it *failed* (either no classroom or constraint conflict),
                         # add it back to the `dynamic_unscheduled_assignments` list.
                         # If it was the `suggested_assign_id` from the template (and wasn't in the initial unscheduled_pool),
                         # and it failed, we don't necessarily need to add it to the dynamic pool, it will be tried again
                         # in the next week's template iteration *if* it still needs sessions.
                         # The simplest reliable method: if attempt failed, don't decrement sessions.
                         # The task remains in the set of tasks needing sessions, which the pool is derived from.
                         # Re-shuffling the pool periodically might be better.

                         # Let's revert to the simpler handling: If the attempted assignment came from the `dynamic_unscheduled_assignments` list (meaning it wasn't the template suggestion, or was a template suggestion that *also* was in the initial unscheduled pool), add it back if it failed.
                         # This requires tracking *how* we got assignment_to_attempt_id.
                         # A potentially simpler approach is to just pick from the dynamic pool always,
                         # prioritizing based on the initial template? That complicates pool management.

                         # Let's go back to the original logic's intent: template provides a *primary* try. If that fails or is done, try the pool.
                         # If the assignment came *from the unscheduled_pool* (`dynamic_unscheduled_assignments`) for this attempt, AND IT FAILED, put it back into `dynamic_unscheduled_assignments`.
                         # If the assignment came from the *template suggestion* (`suggested_assign_id`) AND IT FAILED, it *might* be picked up by the dynamic pool later if it wasn't in the initial unscheduled pool. Or it will be retried next week in the same template slot. This seems okay.

                         # How to know if it came from the dynamic pool? We removed it.
                         # Okay, let's just put it back if it was originally in the pool or was the template suggestion.
                         # Simpler Rule: If assignment_to_attempt_id failed to schedule, add it back to dynamic_unscheduled_assignments if it's not already there and still needs sessions.
                         if assignment_sessions_remaining.get(assignment_to_attempt_id, 0) > 0 and assignment_to_attempt_id not in dynamic_unscheduled_assignments:
                              dynamic_unscheduled_assignments.append(assignment_to_attempt_id)
                              random.shuffle(dynamic_unscheduled_assignments) # Reshuffle after adding back

                 else:
                     # No suitable classroom found, log it
                     conflicts_log.append(
                         {'major_id': assignment.major_id, 'week': week, 'day': day_str, 'period': period_num,
                          'assignment_id': assignment_to_attempt_id,
                          'reason': f"找不到容量({assignment.expected_students})教室"})
                     # print(f"  NOCLASSROOM: Week {week}, {day_str}-{period_num} attempt {assignment_to_attempt_id} failed: No suitable classroom.")
                     # If the attempted assignment came from the dynamic pool and failed, put it back
                     if assignment_sessions_remaining.get(assignment_to_attempt_id, 0) > 0 and assignment_to_attempt_id not in dynamic_unscheduled_assignments:
                           dynamic_unscheduled_assignments.append(assignment_to_attempt_id)
                           random.shuffle(dynamic_unscheduled_assignments) # Reshuffle after adding back


        # End of week loop through timeslots

        # Optional: Re-add any tasks that might have been skipped this week but still need sessions
        # This could happen if a task was only eligible for certain slots (e.g. lab only Fri afternoon)
        # and those specific slots were blocked by global conflicts or teacher preferences *for this week*.
        # They should still be available for future weeks.
        # The dynamic pool approach already handles this implicitly as tasks only get removed permanently when sessions hit 0.
        # Re-shuffling the pool at the start/end of the week loop can help prevent getting stuck.
        random.shuffle(dynamic_unscheduled_assignments)


    # End of week loop

    # Final check for unscheduled tasks
    unscheduled_final = []
    for assign_id, remaining in assignment_sessions_remaining.items():
        if remaining > 0:
            assign_obj = assignments_for_major.get(assign_id)
            course_name, teacher_name = '?', '?'
            if assign_obj:
                course = all_data['courses'].get(assign_obj.course_id)
                teacher = all_data['teachers'].get(assign_obj.teacher_id)
                if course: course_name = course.name
                if teacher: teacher_name = teacher.name
            unscheduled_final.append(
                {'assignment_id': assign_id, 'course_name': course_name, 'teacher_name': teacher_name,
                 'remaining_sessions': remaining})

    # print(f"SCHEDULER: ===== 专业 '{current_major.name}' 排课完成。生成课表条目: {len(final_schedule)}, 冲突记录: {len(conflicts_log)}, 未完成任务数: {len(unscheduled_final)} =====")
    return {'schedule': final_schedule, 'unscheduled_details': unscheduled_final, 'conflicts': conflicts_log}


# ==================================
# 7. 导出到 Excel 函数 (保持不变)
# ==================================
# ... (你的 generate_excel_report_for_send_file 函数保持不变) ...
def sanitize_sheet_name(name):
    name = re.sub(r'[\\/*?:"<>|\[\]]', '_', name)
    return name[:31]


def format_semester_sheet_for_export(worksheet, total_weeks, periods_count):
    if not OPENPYXL_AVAILABLE: return
    try:
        center_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        thin_border_side = Side(border_style="thin", color="000000")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side,
                             bottom=thin_border_side)
        header_font = Font(bold=True)

        for col_idx, column_cells in enumerate(worksheet.columns):
            if col_idx < 2: continue
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if cell.value:
                    try:
                        lines = str(cell.value).split('\n')
                        # Heuristic: Chinese chars approx 1.8 width, others 1
                        cell_len = max(
                            sum(1.8 if '\u4e00' <= char <= '\u9fff' else 1 for char in line) for line in lines)
                        if cell_len > max_length: max_length = cell_len
                    except:
                        pass
            adjusted_width = max_length + 4
            worksheet.column_dimensions[column_letter].width = max(adjusted_width, 15)

        worksheet.column_dimensions['A'].width = 8  # 周数
        worksheet.column_dimensions['B'].width = 8  # 节次

        header_rows = 1
        for row_idx in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 45 # Default row height
            for cell in worksheet[row_idx]:
                cell.alignment = center_alignment
                cell.border = thin_border
            if row_idx == header_rows:
                for cell in worksheet[row_idx]:
                    cell.font = header_font

        # Merge '周数' column cells
        if total_weeks > 0 and periods_count > 0:
            for week in range(1, total_weeks + 1):
                # Find the actual start and end rows for this week based on period count
                # Assuming periods are consecutive 1, 2, 3... for simplicity matching index
                start_row = header_rows + (week - 1) * periods_count + 1
                end_row = start_row + periods_count - 1

                # Ensure the range is within worksheet bounds
                if start_row <= worksheet.max_row and end_row <= worksheet.max_row:
                    try:
                        # Check if cells are not already merged to avoid error
                        # A simple check is to see if the value is already set (from a previous merge attempt or manual edit)
                        # This isn't foolproof but helps prevent common errors.
                        # A better check would be to inspect merge_cells ranges, but that's more complex.
                        # Let's just try and catch the exception if it fails due to existing merge.
                         worksheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                         merged_cell = worksheet.cell(row=start_row, column=1)
                         merged_cell.value = f"第 {week} 周"
                         merged_cell.alignment = center_alignment
                         merged_cell.font = header_font
                         merged_cell.border = thin_border
                    except Exception as merge_e:
                         # print(f"SCHEDULER: 警告：合并单元格 (周数 {week}) 时出错: {merge_e}")
                         # If merge fails, at least set the value in the first cell
                         worksheet.cell(row=start_row, column=1).value = f"第 {week} 周"


    except Exception as format_e:
        print(f"SCHEDULER: 警告：在调整工作表 '{worksheet.title}' 格式时发生错误: {format_e}")
        # import traceback; traceback.print_exc() # Debugging help


def generate_excel_report_for_send_file(schedule_entries, all_data, semester,
                                        target_major_id=None, target_teacher_id=None):
    if not OPENPYXL_AVAILABLE:
        # print("SCHEDULER: 错误：缺少 openpyxl 库，无法导出 Excel。")
        # Return an empty buffer or raise a specific error
        # Raising error is often better as it signals to the caller that the operation failed
        raise ImportError("缺少 openpyxl 库，无法导出 Excel 课表。")

    if not schedule_entries and not (target_major_id or target_teacher_id):
         # If generating a full report but no entries, return empty buffer
         # print("SCHEDULER: 没有排课数据可导出学期总报告。")
         return io.BytesIO()
    elif not schedule_entries and (target_major_id or target_teacher_id):
         # If filtering and no entries match the filter
         # print(f"SCHEDULER: 没有找到符合筛选条件 (Major: {target_major_id}, Teacher: {target_teacher_id}) 的排课数据。")
         # Create an empty sheet indicating no data found
         output_buffer_empty = io.BytesIO()
         try:
              with pd.ExcelWriter(output_buffer_empty, engine='openpyxl') as writer_empty:
                   message = "没有找到相关排课数据。"
                   if target_major_id:
                       major_info = all_data['majors'].get(target_major_id)
                       major_name = major_info.name if major_info else f"专业ID_{target_major_id}"
                       message = f"专业 '{major_name}' 没有排课数据。"
                   elif target_teacher_id:
                        teacher_info = all_data['teachers'].get(target_teacher_id)
                        teacher_name = teacher_info.name if teacher_info else f"教师ID_{target_teacher_id}"
                        message = f"教师 '{teacher_name}' 没有排课数据。"

                   df_empty = pd.DataFrame([[message]])
                   df_empty.to_excel(writer_empty, sheet_name="无数据", index=False, header=False)
              output_buffer_empty.seek(0)
         except Exception as empty_excel_e:
              print(f"SCHEDULER: 生成空 Excel 文件时出错: {empty_excel_e}")
              # Return an empty buffer as a fallback
              return io.BytesIO()

         return output_buffer_empty


    # print(f"SCHEDULER: 开始生成 Excel 报告...")
    output_buffer = io.BytesIO()
    try:
        # Use 'with' to ensure writer is closed properly
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            timeslots = sorted(all_data['timeslots'].values(), key=lambda t: (t.day_of_week, t.period))
            day_map = {'周一': 1, '周二': 2, '周三': 3, '周四': 4, '周五': 5, '周六': 6, '周日': 7}
            periods = sorted(list(set(t.period for t in timeslots)))
            days = sorted(list(set(t.day_of_week for t in timeslots)), key=lambda d: day_map.get(d, 8))
            total_weeks = semester.total_weeks

            if total_weeks <= 0:
                print("SCHEDULER: 错误：学期总周数无效，无法生成学期课表。")
                # Create an error sheet
                df_error = pd.DataFrame([["学期总周数无效，无法生成课表"]])
                df_error.to_excel(writer, sheet_name="错误", index=False, header=False)
                # Need to call close/save implicitly via 'with' or explicitly writer.close()
                # writer.save() # For older pandas
                output_buffer.seek(0)
                return output_buffer


            periods_count = len(periods)
            # Ensure days has all days defined in time slots, even if no classes on them
            all_defined_days = sorted(list(set(ts.day_of_week for ts in all_data['timeslots'].values())), key=lambda d: day_map.get(d, 8))
            if len(days) != len(all_defined_days): # Just a sanity check
                 days = all_defined_days


            # Filter entries if target_major_id or target_teacher_id is provided
            filtered_entries = schedule_entries
            # report_type = "学期总" # Unused variable, removed

            if target_major_id:
                filtered_entries = [e for e in schedule_entries if e.major_id == target_major_id]
                if not filtered_entries:
                    # print(f"SCHEDULER: 没有找到专业 ID {target_major_id} 的排课数据。")
                    # If filter yields no results, maybe write an empty sheet or handle upstream
                    pass # Handle this case before calling this function or write an empty sheet here

                major_info = all_data['majors'].get(target_major_id)
                entity_name = major_info.name if major_info else f"专业ID_{target_major_id}"

                # Generate single sheet for the major
                df_major_semester = pd.DataFrame(
                    index=pd.MultiIndex.from_product([range(1, total_weeks + 1), periods], names=['周数', '节次']),
                    columns=days
                ).fillna('')
                for entry in filtered_entries:
                    ts = all_data['timeslots'].get(entry.timeslot_id)
                    if not ts or ts.day_of_week not in days or ts.period not in periods:
                         # print(f"SCHEDULER: 警告：排课条目 {entry.id} 的时间段 {entry.timeslot_id} 数据异常，跳过导出。")
                         continue # Skip entry if timeslot data is inconsistent

                    course = all_data['courses'].get(entry.course_id)
                    teacher = all_data['teachers'].get(entry.teacher_id)
                    classroom = all_data['classrooms'].get(entry.classroom_id)
                    cell_text = f"{course.name if course else '?'}\n{teacher.name if teacher else '?'}\n@{classroom.name if classroom else '?'}"
                    # Make sure day_of_week and period exist in the DataFrame index/columns
                    if (entry.week_number, ts.period) in df_major_semester.index and ts.day_of_week in df_major_semester.columns:
                         df_major_semester.loc[(entry.week_number, ts.period), ts.day_of_week] = cell_text
                    # else:
                    #      print(f"SCHEDULER: 警告：排课条目 {entry.id} 的周/节次/日期 ({entry.week_number}, {ts.period}, {ts.day_of_week}) 与DataFrame索引/列不匹配，跳过导出。")


                sheet_name_major = sanitize_sheet_name(f"专业_{entity_name}")
                df_major_semester.to_excel(writer, sheet_name=sheet_name_major, merge_cells=False)
                worksheet_major = writer.sheets[sheet_name_major]
                format_semester_sheet_for_export(worksheet_major, total_weeks, periods_count)


            elif target_teacher_id:
                filtered_entries = [e for e in schedule_entries if e.teacher_id == target_teacher_id]
                if not filtered_entries:
                     # print(f"SCHEDULER: 没有找到教师 ID {target_teacher_id} 的排课数据。")
                     pass # Handle upstream or write empty sheet

                teacher_info = all_data['teachers'].get(target_teacher_id)
                entity_name = teacher_info.name if teacher_info else f"教师ID_{target_teacher_id}"

                df_teacher_semester = pd.DataFrame(
                    index=pd.MultiIndex.from_product([range(1, total_weeks + 1), periods], names=['周数', '节次']),
                    columns=days
                ).fillna('')
                for entry in filtered_entries:
                    ts = all_data['timeslots'].get(entry.timeslot_id)
                    if not ts or ts.day_of_week not in days or ts.period not in periods:
                         # print(f"SCHEDULER: 警告：排课条目 {entry.id} 的时间段 {entry.timeslot_id} 数据异常，跳过导出。")
                         continue

                    course = all_data['courses'].get(entry.course_id)
                    major = all_data['majors'].get(entry.major_id)
                    classroom = all_data['classrooms'].get(entry.classroom_id)
                    cell_text = f"{course.name if course else '?'}\n({major.name if major else '?'})\n@{classroom.name if classroom else '?'}"
                    # Make sure day_of_week and period exist in the DataFrame index/columns
                    if (entry.week_number, ts.period) in df_teacher_semester.index and ts.day_of_week in df_teacher_semester.columns:
                         df_teacher_semester.loc[(entry.week_number, ts.period), ts.day_of_week] = cell_text
                    # else:
                    #      print(f"SCHEDULER: 警告：排课条目 {entry.id} 的周/节次/日期 ({entry.week_number}, {ts.period}, {ts.day_of_week}) 与DataFrame索引/列不匹配，跳过导出。")


                sheet_name_teacher = sanitize_sheet_name(f"教师_{entity_name}")
                df_teacher_semester.to_excel(writer, sheet_name=sheet_name_teacher, merge_cells=False)
                worksheet_teacher = writer.sheets[sheet_name_teacher]
                format_semester_sheet_for_export(worksheet_teacher, total_weeks, periods_count)

            else:  # Full semester report (all majors and all teachers)
                # A. 专业课表
                schedule_by_major = defaultdict(list)
                # Only include majors that actually have scheduled entries
                involved_major_ids = sorted(list(set(entry.major_id for entry in filtered_entries)), key=lambda mid: all_data['majors'].get(mid, Major(id=mid, name=f"UnknownMajor{mid}")).name)
                for entry in filtered_entries: schedule_by_major[entry.major_id].append(entry)

                for major_id_iter in involved_major_ids:
                    major_schedule = schedule_by_major[major_id_iter]
                    major_info_iter = all_data['majors'].get(major_id_iter)
                    major_name_iter = major_info_iter.name if major_info_iter else f"专业ID_{major_id_iter}"
                    df_major_semester = pd.DataFrame(
                        index=pd.MultiIndex.from_product([range(1, total_weeks + 1), periods], names=['周数', '节次']),
                        columns=days
                    ).fillna('')
                    for entry in major_schedule:  # Use major_schedule, not filtered_entries
                        ts = all_data['timeslots'].get(entry.timeslot_id)
                        if not ts or ts.day_of_week not in days or ts.period not in periods:
                             # print(f"SCHEDULER: 警告：排课条目 {entry.id} 的时间段 {entry.timeslot_id} 数据异常，跳过导出专业课表。")
                             continue
                        course = all_data['courses'].get(entry.course_id)
                        teacher = all_data['teachers'].get(entry.teacher_id)
                        classroom = all_data['classrooms'].get(entry.classroom_id)
                        cell_text = f"{course.name if course else '?'}\n{teacher.name if teacher else '?'}\n@{classroom.name if classroom else '?'}"
                        if (entry.week_number, ts.period) in df_major_semester.index and ts.day_of_week in df_major_semester.columns:
                            df_major_semester.loc[(entry.week_number, ts.period), ts.day_of_week] = cell_text
                        # else:
                        #      print(f"SCHEDULER: 警告：排课条目 {entry.id} 的周/节次/日期 ({entry.week_number}, {ts.period}, {ts.day_of_week}) 与DataFrame索引/列不匹配，跳过导出专业课表。")


                    sheet_name_major = sanitize_sheet_name(f"专业_{major_name_iter}")
                    df_major_semester.to_excel(writer, sheet_name=sheet_name_major, merge_cells=False)
                    worksheet_major = writer.sheets[sheet_name_major]
                    format_semester_sheet_for_export(worksheet_major, total_weeks, periods_count)

                # B. 教师课表
                schedule_by_teacher = defaultdict(list)
                 # Only include teachers that actually have scheduled entries
                involved_teacher_ids = sorted(list(set(entry.teacher_id for entry in filtered_entries)), key=lambda tid: all_data['teachers'].get(tid, Teacher(id=tid, user_id=None, name=f"UnknownTeacher{tid}")).name)
                for entry in filtered_entries: schedule_by_teacher[entry.teacher_id].append(entry)

                for teacher_id_iter in involved_teacher_ids:
                    teacher_schedule = schedule_by_teacher[teacher_id_iter]
                    teacher_info_iter = all_data['teachers'].get(teacher_id_iter)
                    teacher_name_iter = teacher_info_iter.name if teacher_info_iter else f"教师ID_{teacher_id_iter}"
                    df_teacher_semester = pd.DataFrame(
                        index=pd.MultiIndex.from_product([range(1, total_weeks + 1), periods], names=['周数', '节次']),
                        columns=days
                    ).fillna('')
                    for entry in teacher_schedule:  # Use teacher_schedule
                        ts = all_data['timeslots'].get(entry.timeslot_id)
                        if not ts or ts.day_of_week not in days or ts.period not in periods:
                             # print(f"SCHEDULER: 警告：排课条目 {entry.id} 的时间段 {entry.timeslot_id} 数据异常，跳过导出教师课表。")
                             continue
                        course = all_data['courses'].get(entry.course_id)
                        major = all_data['majors'].get(entry.major_id)
                        classroom = all_data['classrooms'].get(entry.classroom_id)
                        cell_text = f"{course.name if course else '?'}\n({major.name if major else '?'})\n@{classroom.name if classroom else '?'}"
                        if (entry.week_number, ts.period) in df_teacher_semester.index and ts.day_of_week in df_teacher_semester.columns:
                            df_teacher_semester.loc[(entry.week_number, ts.period), ts.day_of_week] = cell_text
                        # else:
                        #      print(f"SCHEDULER: 警告：排课条目 {entry.id} 的周/节次/日期 ({entry.week_number}, {ts.period}, {ts.day_of_week}) 与DataFrame索引/列不匹配，跳过导出教师课表。")

                    sheet_name_teacher = sanitize_sheet_name(f"教师_{teacher_name_iter}")
                    df_teacher_semester.to_excel(writer, sheet_name=sheet_name_teacher, merge_cells=False)
                    worksheet_teacher = writer.sheets[sheet_name_teacher]
                    format_semester_sheet_for_export(worksheet_teacher, total_weeks, periods_count)

        # writer.save() # For older pandas version. For newer, 'with' handles it.
        output_buffer.seek(0)
        # print(f"SCHEDULER: Excel 报告生成完毕。")
        return output_buffer

    except Exception as e:
        print(f"SCHEDULER: 生成 Excel 文件时出错: {e}")
        # import traceback; traceback.print_exc()
        # Return an empty buffer or re-raise to indicate failure
        # For now, returning an empty buffer to avoid breaking send_file if it expects BytesIO
        error_output = io.BytesIO()
        try:
             with pd.ExcelWriter(error_output, engine='openpyxl') as writer_error:
                 pd.DataFrame([["生成Excel时出错:", str(e)]]).to_excel(writer_error, sheet_name="错误", index=False, header=False)
             error_output.seek(0)
        except Exception as inner_e:
             print(f"SCHEDULER: 写入Excel错误信息时也出错: {inner_e}")
             return io.BytesIO() # Return empty buffer if even error logging fails

        return error_output # Return buffer with error message


# ==================================
# 8. 数据库操作函数 (保持不变)
# ==================================
# ... (你的 clear_db_for_semester 和 save_schedule_to_db 函数保持不变) ...
def clear_db_for_semester(semester_id, get_connection_func):
    conn = None
    cur = None
    try:
        conn = get_connection_func()
        cur = conn.cursor()
        # print(f"SCHEDULER: 正在清空数据库中 学期 ID={semester_id} 的旧排课记录...")
        delete_query = "DELETE FROM timetable_entries WHERE semester_id = %s"
        cur.execute(delete_query, (semester_id,))
        deleted_count = cur.rowcount
        conn.commit()
        cur.close()
        # print(f"SCHEDULER: 成功删除 {deleted_count} 条旧记录。")
        return True, deleted_count
    except psycopg2.Error as e:
        print(f"SCHEDULER: 清空学期 {semester_id} 的数据库记录时出错: {e}")
        if conn: conn.rollback()
        # import traceback; traceback.print_exc()
        raise  # Re-raise
    except Exception as e:
        print(f"SCHEDULER: 清空记录过程中发生未知错误: {e}")
        if conn: conn.rollback()
        raise
    finally:
        if cur: cur.close()
        if conn: conn.close()


def save_schedule_to_db(schedule_entries, get_connection_func):
    if not schedule_entries:
        # print("SCHEDULER: 没有排课条目需要保存。")
        return 0
    conn_save = None
    cur_save = None
    inserted_count = 0
    try:
        conn_save = get_connection_func()
        cur_save = conn_save.cursor()

        # insert_query 应该在 VALUES 后面只有一个 %s
        insert_query = """
            INSERT INTO timetable_entries
            (semester_id, major_id, course_id, teacher_id, classroom_id, timeslot_id, week_number, assignment_id)
            VALUES %s
        """

        data_to_insert = [
            (e.semester_id, e.major_id, e.course_id, e.teacher_id, e.classroom_id, e.timeslot_id, e.week_number,
             e.assignment_id)
            for e in schedule_entries
        ]
        # print("[DEBUG] 插入语句 (用于 execute_values):", insert_query)
        # print("[DEBUG] 首条数据:", data_to_insert[0] if data_to_insert else "无数据")

        # Use execute_values for efficient batch insertion
        from psycopg2.extras import execute_values

        # Default page_size is 100, increase for better performance with many rows
        # Or set to len(data_to_insert) if dataset size is manageable
        execute_values(cur_save, insert_query, data_to_insert, page_size=max(100, len(data_to_insert)//10))

        inserted_count = len(data_to_insert) # Assume all attempted rows are inserted if no error
        conn_save.commit()

        cur_save.close()
        # print(f"SCHEDULER: 成功保存 {inserted_count} 条排课记录到数据库。")
        return inserted_count
    except psycopg2.Error as e:
        print(f"SCHEDULER: 保存排课结果到数据库时出错: {e}")
        if conn_save: conn_save.rollback()
        # import traceback; traceback.print_exc()
        raise  # Re-raise
    except Exception as e:
         print(f"SCHEDULER: 保存记录过程中发生未知错误: {e}")
         if conn_save: conn_save.rollback()
         raise
    finally:
        if cur_save: cur_save.close()
        if conn_save: conn_save.close()


# ==================================
# 9. 主排课流程函数 (保持不变，因为它已经正确传递全局状态)
# ==================================
from collections import defaultdict
import random
# Assume necessary classes (Course, Major, etc.) and functions are defined elsewhere and correctly imported.
# Assume get_connection_func returns a standard DB-API 2 connection object.

# IMPORTANT: You need to replace 'processed' with the actual status value (e.g., integer code, string)
# that you use in your database schema for teacher_preferences to mark them as processed.

def run_full_scheduling_process(target_semester_id, get_connection_func):
    """
    主排课流程函数，被 Flask API 调用。
    返回一个包含排课结果摘要的字典。
    在排课完成后（无论成功或失败）尝试更新所有教师偏好状态。
    """
    print(f"SCHEDULER: 开始执行学期 ID {target_semester_id} 的自动排课程序...")
    summary = {
        "status": "failure",
        "message": "",
        "processed_majors": 0,
        "total_scheduled_entries": 0,
        "total_conflicts": 0,
        "total_uncompleted_tasks": 0,
        "db_records_cleared": 0,
        "db_records_saved": 0,
        "details": []  # For per-major messages or errors
    }

    # Initialize all_data and all_assignments_in_semester outside try for potential use in finally (though simplified logic in finally might not need them)
    all_data = None
    all_assignments_in_semester = defaultdict(dict)

    try:
        all_data = load_data_from_db(get_connection_func)
        if not all_data:
            summary["message"] = "数据加载失败。"
            # Note: The finally block will still attempt to update preferences even if data loading fails.
            return summary

        current_semester = all_data['semesters'].get(target_semester_id)
        if not current_semester:
            summary["message"] = f"未找到 ID 为 {target_semester_id} 的学期信息。"
            return summary
        if current_semester.total_weeks <= 0:
             summary["message"] = f"目标学期 '{current_semester.name}' (ID: {target_semester_id}) 总周数 ({current_semester.total_weeks}) 无效。"
             return summary

        majors_in_semester = set()
        # Populate all_assignments_in_semester and find majors involved in this semester
        for assign_id, assign in all_data['course_assignments'].items():
            if assign.semester_id == target_semester_id:
                majors_in_semester.add(assign.major_id)
                all_assignments_in_semester[assign.major_id][assign_id] = assign # Populate the defaultdict

        if not majors_in_semester:
            summary["message"] = f"学期 '{current_semester.name}' (ID: {target_semester_id}) 中未找到任何专业的教学任务。"
            summary["status"] = "success_no_tasks"  # Special status
            # Note: The finally block will still attempt to update preferences even if no tasks are found.
            return summary

        # 清空旧记录
        clear_success, cleared_count = clear_db_for_semester(target_semester_id, get_connection_func)
        summary["db_records_cleared"] = cleared_count
        if not clear_success:
             summary["message"] = "清空旧排课记录失败。"
             summary["status"] = "error"
             return summary

        all_final_schedule_entries_for_semester = []

        # --- 在这里初始化一次全局状态 ---
        master_global_timetable_state = {'teacher_schedule': set(), 'classroom_schedule': set(),
                                         'major_schedule': set()}
        # --- 初始化结束 ---

        # Sort majors for consistent processing order
        # Using a lambda that safely gets the name, providing a default object structure if Major not found
        # This sorting fix might still be needed if Major() takes no arguments in your code
        sorted_major_ids = sorted(
            list(majors_in_semester),
            key=lambda mid: all_data['majors'].get(
                mid,
                # Create a simple object with a 'name' attribute for sorting purposes
                type('MajorSortHelper', (object,), {'name': f"未知专业ID_{mid}"})()
            ).name
        )


        # --- Major scheduling loop ---
        for major_id in sorted_major_ids:
            current_major = all_data['majors'].get(major_id)
            assignments_for_this_major = all_assignments_in_semester.get(major_id, {})
            major_name = current_major.name if current_major else f"未知专业ID_{major_id}"

            major_detail_msg = f"专业 '{major_name}' (ID: {major_id}): "
            if not assignments_for_this_major:
                major_detail_msg += "没有教学任务，跳过。"
                summary["details"].append(major_detail_msg)
                continue

            # Assuming generate_initial_template and schedule_with_generated_template are defined elsewhere
            # Ensure generate_initial_template has the day_order fix if needed
            initial_template_dp, unscheduled_pool = generate_initial_template(assignments_for_this_major, all_data)

            # Pass Major object or a dummy object with id and name if current_major is None
            major_obj_for_scheduling = current_major if current_major else type('MajorDummy', (object,), {'id': major_id, 'name': major_name})()

            schedule_result_obj = schedule_with_generated_template(
                assignments_for_this_major, current_semester,
                major_obj_for_scheduling, # Pass the major object or dummy
                all_data, initial_template_dp, unscheduled_pool,
                master_global_timetable_state  # <-- 传递全局状态
            )

            major_schedule = schedule_result_obj.get('schedule', [])
            all_final_schedule_entries_for_semester.extend(major_schedule)

            # master_global_timetable_state is updated inside schedule_with_generated_template

            num_scheduled_major = len(major_schedule)
            major_conflicts_log = schedule_result_obj.get('conflicts', [])
            num_conflicts_major = len(major_conflicts_log)
            major_unscheduled_details = schedule_result_obj.get('unscheduled_details', [])
            num_uncompleted_major = len(major_unscheduled_details)

            summary["processed_majors"] += 1
            summary["total_scheduled_entries"] += num_scheduled_major
            summary["total_conflicts"] += num_conflicts_major
            summary["total_uncompleted_tasks"] += num_uncompleted_major

            major_detail_msg += f"生成课表 {num_scheduled_major}条, 冲突 {num_conflicts_major}次, 未完成任务 {num_uncompleted_major}个。"
            summary["details"].append(major_detail_msg)
        # --- End of Major scheduling loop ---


        # Save all results for the semester at once
        if all_final_schedule_entries_for_semester:
            saved_count_total = save_schedule_to_db(all_final_schedule_entries_for_semester, get_connection_func)
            summary["db_records_saved"] = saved_count_total


        summary["status"] = "success"
        summary["message"] = f"学期 {target_semester_id} 排课完成。"
        if summary["total_conflicts"] > 0 or summary["total_uncompleted_tasks"] > 0:
            summary["message"] += f" 总冲突: {summary['total_conflicts']}次, 未完成任务: {summary['total_uncompleted_tasks']}个。"

        # The return happens here if no exception occurs,
        # but the finally block will execute regardless.

    except Exception as e:
        print(f"SCHEDULER: 排课主流程发生严重错误: {e}")
        # import traceback; traceback.print_exc() # Uncomment for detailed traceback during debugging
        summary["message"] = f"排课过程中发生错误: {str(e)}"
        summary["status"] = "error"
        # No explicit rollback needed here if sub-functions handle their transactions or if connection is not shared globally.
        # The finally block will execute after this exception handler finishes.

    finally:
        # --- START: Update ALL teacher preference status ---
        # This block executes after try/except, regardless of whether an error occurred or a return happened.
        print("SCHEDULER: 排课流程结束，尝试更新所有教师偏好状态...")
        update_conn = None
        update_cursor = None
        try:
            update_conn = get_connection_func() # Get a fresh connection for the update
            update_cursor = update_conn.cursor()

            # --- IMPORTANT: Replace 'processed' with your actual status value/logic ---
            # This UPDATE statement targets ALL rows in the 'teacher_preferences' table.
            # If you need to filter (e.g., by semester, initial status), add a WHERE clause.
            new_status_value = "applied" # <--- REPLACE THIS WITH YOUR ACTUAL STATUS VALUE

            # NOTE: Table name 'teacher_preferences' must match your DB schema
            update_query = "UPDATE teacher_scheduling_preferences SET status = 'applied'"

            # Execute the query
            update_cursor.execute(update_query) # Pass the status value as a single-item tuple
            update_conn.commit() # Commit the changes

            # You might want to fetch sqlca.sqlerrd[2] or similar for rows affected, but DB-API doesn't guarantee it.
            # A simple print indicates the attempt was made.
            print(f"SCHEDULER: 已尝试更新数据库中所有教师偏好状态为 '{new_status_value}'。请检查数据库确认。")

        except Exception as update_e:
            print(f"SCHEDULER: 在 finally 块中更新教师偏好状态时发生错误: {update_e}")
            # import traceback; traceback.print_exc() # Uncomment for detailed traceback
            if update_conn:
                try:
                    update_conn.rollback() # Attempt to rollback the update transaction
                    print("SCHEDULER: 教师偏好状态更新事务已回滚。")
                except Exception as rb_e:
                    print(f"SCHEDULER: 回滚教师偏好状态更新事务时发生错误: {rb_e}")

        finally:
            # Ensure cursor and connection are closed
            if update_cursor:
                try: update_cursor.close()
                except: pass # Ignore close errors
            if update_conn:
                try: update_conn.close()
                except: pass # Ignore close errors
            # --- END: Update ALL teacher preference status ---

        # The final return value is the 'summary' dictionary, which holds the result of the scheduling attempt.
        return summary

# Assume other necessary function and class definitions are present here or imported.
